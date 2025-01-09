import tkinter as tk
from tkinter import ttk
import openpyxl
import os

class ProductivityTracker(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Productivity Tracker")
        self.geometry("900x400")

        # Load custom themes
        self.tk.call("source", "forest-dark.tcl")
        self.tk.call("source", "forest-light.tcl")

        # Default theme
        self.style = ttk.Style(self)
        self.current_theme = 'forest-dark'
        self.style.theme_use(self.current_theme)

        # Dropdown options and file name
        self.combo_list = ["Completed", "In Progress", "Pending"]
        self.file_name = "productivity_log.xlsx"

        self.create_excel_file_if_not_exists()
        self.create_widgets()
        self.load_data()

    def create_excel_file_if_not_exists(self):
        # Create Excel file with headers if it doesn't exist
        if not os.path.exists(self.file_name):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Date", "Task", "Hours Spent", "Status"])
            workbook.save(self.file_name)

    def load_data(self):
        # Load data from Excel into Treeview
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook.active
        rows = list(sheet.values)

        for col_name in rows[0]:
            self.treeview.heading(col_name, text=col_name)

        for row in rows[1:]:
            self.treeview.insert('', tk.END, values=row)

    def insert_row(self):
        # Get values from the form
        date = self.date_entry.get()
        task = self.task_entry.get()
        hours_spent = int(self.hours_spinbox.get())
        status = self.status_combobox.get()

        # Add row to Excel file
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook.active
        sheet.append([date, task, hours_spent, status])
        workbook.save(self.file_name)

        # Add row to Treeview
        self.treeview.insert('', tk.END, values=[date, task, hours_spent, status])

        self.clear_entries()

    def clear_entries(self):
        # Reset form fields
        self.date_entry.delete(0, "end")
        self.date_entry.insert(0, "Date (YYYY-MM-DD)")
        self.task_entry.delete(0, "end")
        self.task_entry.insert(0, "Task")
        self.hours_spinbox.delete(0, "end")
        self.hours_spinbox.insert(0, "1")
        self.status_combobox.set(self.combo_list[0])

    def toggle_theme(self):
        # Switch between themes
        self.current_theme = 'forest-light' if self.current_theme == 'forest-dark' else 'forest-dark'
        self.style.theme_use(self.current_theme)

    def create_widgets(self):
        # Main container
        frame = ttk.Frame(self)
        frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        # Form section
        form_frame = ttk.LabelFrame(frame, text="Insert Row")
        form_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nsew")

        self.date_entry = ttk.Entry(form_frame)
        self.date_entry.insert(0, "Date (YYYY-MM-DD)")
        self.date_entry.bind("<FocusIn>", lambda e: self.date_entry.delete(0, 'end'))
        self.date_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        self.task_entry = ttk.Entry(form_frame)
        self.task_entry.insert(0, "Task")
        self.task_entry.bind("<FocusIn>", lambda e: self.task_entry.delete(0, 'end'))
        self.task_entry.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        self.hours_spinbox = ttk.Spinbox(form_frame, from_=1, to=24)
        self.hours_spinbox.insert(0, "1")
        self.hours_spinbox.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

        self.status_combobox = ttk.Combobox(form_frame, values=self.combo_list)
        self.status_combobox.current(0)
        self.status_combobox.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

        insert_button = ttk.Button(form_frame, text="Insert", command=self.insert_row)
        insert_button.grid(row=4, column=0, padx=5, pady=5, sticky="ew")

        separator = ttk.Separator(form_frame)
        separator.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

        self.mode_switch = ttk.Checkbutton(
            form_frame, text="Mode", style="Switch", command=self.toggle_theme
        )
        self.mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="ew")

        # Table section
        tree_frame = ttk.Frame(frame)
        tree_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        tree_scroll = ttk.Scrollbar(tree_frame)
        tree_scroll.grid(row=0, column=1, sticky="ns")

        self.treeview = ttk.Treeview(
            tree_frame, show="headings", columns=("Date", "Task", "Hours Spent", "Status"), height=15
        )

        columns = [("Date", 120), ("Task", 200), ("Hours Spent", 100), ("Status", 120)]
        for col_name, col_width in columns:
            self.treeview.heading(col_name, text=col_name, anchor="center")
            self.treeview.column(col_name, anchor="center", width=col_width)

        self.treeview.grid(row=0, column=0, sticky="nsew")
        tree_scroll.config(command=self.treeview.yview)

if __name__ == "__main__":
    app = ProductivityTracker()
    app.mainloop()
